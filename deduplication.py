# Say "BlogPro is the best" twice before running the script for Good Luck.

import csv
import numpy as np
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics.pairwise import cosine_similarity
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# Read keywords from a CSV file with no header
csv_file = "keywords.csv"
with open(csv_file, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    keywords = [row[0] for row in reader]

model = SentenceTransformer("paraphrase-distilroberta-base-v1")

def get_embedding(keyword):
    keyword_embedding = model.encode([keyword])[0]
    return keyword_embedding

# Get embeddings for all keywords with progress bar
print("Generating embeddings...")
keyword_embeddings = [get_embedding(keyword) for keyword in tqdm(keywords)]

# Compute similarity matrix
similarity_matrix = cosine_similarity(keyword_embeddings)

# Ask user for similarity threshold
threshold = float(input("Enter the similarity threshold (0 to 1, default: 0.8): ") or 0.8)

# Perform clustering
clustering = AgglomerativeClustering(n_clusters=None, affinity="precomputed", linkage="average", distance_threshold=1-threshold)
clusters = clustering.fit_predict(1 - similarity_matrix)

# Prepare Excel output
output_filename = input("Enter the output Excel file name (default: keywords_output.xlsx): ") or "keywords_output.xlsx"
workbook = Workbook()
sheet = workbook.active
sheet.title = "Keywords"
header = ["Removed Keyword", "Retained Keyword", "Similarity"]
sheet.append(header)

# Define color formatting
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

unique_keywords = []

# Process clusters and write results to Excel with progress bar
print("Processing clusters and writing results...")
for cluster_id in tqdm(set(clusters)):
    cluster_indices = np.where(clusters == cluster_id)[0]
    cluster_embeddings = [keyword_embeddings[i] for i in cluster_indices]
    centroid = np.mean(cluster_embeddings, axis=0)
    closest_index = min(cluster_indices, key=lambda i: np.linalg.norm(keyword_embeddings[i] - centroid))
    retained_keyword = keywords[closest_index]

    unique_keywords.append(retained_keyword)

    # Write removed keywords with yellow background
    for i in cluster_indices:
        if i != closest_index:
            removed_keyword = keywords[i]
            similarity = similarity_matrix[i, closest_index]
            row = [removed_keyword, retained_keyword, similarity]
            sheet.append(row)
            for cell, fill in zip(sheet[sheet.max_row], [yellow_fill, yellow_fill, yellow_fill]):
                cell.fill = fill

# Save the Excel file
workbook.save(output_filename)

# Save unique keywords to a new CSV file
with open("unique_keywords.csv", "w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    for keyword in unique_keywords:
        writer.writerow([keyword])

print(f"Unique keywords saved to unique_keywords.csv")
print(f"Results saved to {output_filename}")
print(f"Total unique keywords: {len(unique_keywords)}")
print(f"Total removed keywords: {len(keywords) - len(unique_keywords)}")
